import json
import logging
import os
import time
import threading
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from kafka_config import (
    DEFAULT_BATCH_SIZE, DEFAULT_LINGER_MS, DEFAULT_COMPRESSION_TYPE,
    DEFAULT_MAX_REQUEST_SIZE, DEFAULT_ACKS
)

# Path to the unified JSON log file and Excel log files
json_log_file_path = "log_files/sys_logs.json"
producer_xlsx_log_file_path = "log_files/producer_logs.xlsx"
consumer_xlsx_log_file_path = "log_files/consumer_logs.xlsx"

# Ensure the log directory exists
os.makedirs(os.path.dirname(json_log_file_path), exist_ok=True)
os.makedirs(os.path.dirname(producer_xlsx_log_file_path), exist_ok=True)
os.makedirs(os.path.dirname(consumer_xlsx_log_file_path), exist_ok=True)

# Set up logging for JSON file and console output
logging.basicConfig(
    level=logging.DEBUG,  # Log all levels
    format='{"timestamp": "%(asctime)s", "level": "%(levelname)s", "thread_id": "%(thread)d", "case_id": "%(thread)d", "entity_name": "%(name)s", "action": "%(message)s"},',
    handlers=[
        logging.FileHandler(json_log_file_path, mode='a'),
        # logging.StreamHandler()
    ]
)

# Create workbooks and sheets for Excel logging
producer_wb = Workbook()
producer_ws = producer_wb.active
producer_ws.title = "Producer Logs"

consumer_wb = Workbook()
consumer_ws = consumer_wb.active
consumer_ws.title = "Consumer Logs"

# Set up the Excel sheet headers for Producer logs
producer_headers = [
    "Timestamp", "Case ID", "Producer ID", "Action", "Batch Size", 
    "Linger MS", "Compression Type", "Max Request Size", "Acks", "Message Details"
]
for col_num, header in enumerate(producer_headers, 1):
    col_letter = get_column_letter(col_num)
    producer_ws[f"{col_letter}1"] = header

# Set up the Excel sheet headers for Consumer logs
consumer_headers = [
    "Timestamp", "Case ID", "Consumer ID", "Action", "Message Details"
]
for col_num, header in enumerate(consumer_headers, 1):
    col_letter = get_column_letter(col_num)
    consumer_ws[f"{col_letter}1"] = header

def log_producer_operation(producer_id, action, message, success=True, buffer = None):
    """
    Log producer actions to a unified JSON log file, Excel file, and console.
    """
    thread_id = threading.get_ident()
    timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
    
    # If success is False, capture the error message (if any)
    message_details = message if success else f"Error: {message}"

    # If message is a dictionary, convert it to JSON string before logging it
    if isinstance(message, dict):
        message_details = json.dumps(message)

    log_entry = {
        "timestamp": timestamp,
        "case_id": thread_id,
        "action": action,
        "producer_id": producer_id,
        "batch_size": DEFAULT_BATCH_SIZE,
        "linger_ms": DEFAULT_LINGER_MS,
        "compression_type": DEFAULT_COMPRESSION_TYPE,
        "max_request_size": DEFAULT_MAX_REQUEST_SIZE,
        "acks": DEFAULT_ACKS,
        "message_details": message_details
    }

    # Log to the JSON file
    logging.info(json.dumps(log_entry))
    if buffer is not None:
        buffer.add_log(log_entry)

    # Log to the Excel file (Producer logs)
    row = [
        timestamp, thread_id, producer_id, action,
        DEFAULT_BATCH_SIZE, DEFAULT_LINGER_MS, DEFAULT_COMPRESSION_TYPE,
        DEFAULT_MAX_REQUEST_SIZE, DEFAULT_ACKS, message_details
    ]
    producer_ws.append(row)
    producer_wb.save(producer_xlsx_log_file_path)

def log_consumer_operation(consumer_id, action, message, success=True, buffer = None):
    """
    Log consumer actions to a unified JSON log file, Excel file, and console.
    """
    thread_id = threading.get_ident()
    timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
    
    # If success is False, capture the error message (if any)
    message_details = message if success else f"Error: {message}"

    # If message is a dictionary, convert it to JSON string before logging it
    if isinstance(message, dict):
        message_details = json.dumps(message)

    log_entry = {
        "timestamp": timestamp,
        "case_id": thread_id,
        "action": action,
        "consumer_id": consumer_id,
        "batch_size": DEFAULT_BATCH_SIZE,
        "linger_ms": DEFAULT_LINGER_MS,
        "compression_type": DEFAULT_COMPRESSION_TYPE,
        "max_request_size": DEFAULT_MAX_REQUEST_SIZE,
        "acks": DEFAULT_ACKS,
        "message_details": message_details
    }

    # Log to the JSON file
    logging.info(json.dumps(log_entry))
    
    if buffer is not None:
        buffer.add_log(log_entry)

    # Log to the Excel file (Consumer logs)
    row = [
        timestamp, thread_id, consumer_id, action, message_details
    ]
    consumer_ws.append(row)
    consumer_wb.save(consumer_xlsx_log_file_path)
